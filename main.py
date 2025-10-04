import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ==============================
# USER INPUT
# ==============================
job_title = input("Enter job title: ").strip().replace(" ", "+")
remote_location = input("Enter remote location: ").strip().lower()

# ==============================
# SELENIUM SETUP
# ==============================
service = Service("chromedriver.exe")  # Path to your ChromeDriver
driver = webdriver.Chrome(service=service)

url = f"https://remoteok.com/remote-{job_title}-jobs"
driver.get(url)

print(f"Loading jobs for '{job_title}'...")
time.sleep(5)  # Wait for JavaScript to load

html = driver.page_source
driver.quit()

# ==============================
# PARSE HTML
# ==============================
soup = BeautifulSoup(html, "html.parser")
job_rows = soup.find_all("tr", {"class": "job", "data-id": True})

jobs = []
for job in job_rows:
    title_tag = job.find("h2", {"itemprop": "title"})
    company_tag = job.find("h3", {"itemprop": "name"})
    location_tag = job.find("div", {"class": "location"})
    link_tag = job.find("a", {"itemprop": "url"})

    title = title_tag.text.strip() if title_tag else "N/A"
    company = company_tag.text.strip() if company_tag else "N/A"
    location = location_tag.text.strip() if location_tag else "N/A"
    link = f"https://remoteok.com{link_tag['href']}" if link_tag and link_tag.has_attr("href") else "N/A"

    jobs.append((title, company, location, link))

print(f"\nTotal jobs found: {len(jobs)}")

# ==============================
# CREATE EXCEL REPORT
# ==============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "RemoteOK Jobs"

# Headers
headers = ["Job Title", "Company", "Location", "Job Link"]
header_font = Font(bold=True)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ws.append(headers)
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = header_font
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

# Add job data
for job in jobs:
    ws.append(job)

# Highlight matching location in green
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
    location_cell = row[2]  # third column = Location
    if remote_location in location_cell.value.lower():
        for cell in row:
            cell.fill = green_fill

# Add total count at bottom
total_row = ws.max_row + 2
ws[f"A{total_row}"] = "Total Jobs:"
ws[f"A{total_row}"].font = Font(bold=True)
ws[f"B{total_row}"] = len(jobs)
ws[f"B{total_row}"].font = Font(bold=True)

# Auto-adjust column widths
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length + 2

# Save Excel file
file_name = f"RemoteOK_{job_title}_Jobs.xlsx"
wb.save(file_name)
print(f"\n✅ Excel report created: {file_name}")
